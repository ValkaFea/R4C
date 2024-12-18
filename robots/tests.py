import datetime
import io
from django.test import TestCase
from django.urls import reverse
import json
from django.utils import timezone
from openpyxl import load_workbook
from .models import Robot
from django.core import mail
from orders.models import Order
from customers.models import Customer

class RobotAPITest(TestCase):
    def test_add_robot_success(self):
        url = reverse('add_robot')
        data = {
            'model': 'R2',
            'version': 'D2',
            'created': '2022-12-31 23:59:59'
        }
        response = self.client.post(url, json.dumps(data), content_type='application/json')
        self.assertEqual(response.status_code, 201)
        self.assertIn('Robot created successfully', response.json()['message'])

    def test_add_robot_missing_fields(self):
        url = reverse('add_robot')
        data = {
            'model': 'R2',
            'created': '2022-12-31 23:59:59'
        }
        response = self.client.post(url, json.dumps(data), content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertIn('Missing required fields', response.json()['error'])

    def test_add_robot_invalid_date(self):
        url = reverse('add_robot')
        data = {
            'model': 'R2',
            'version': 'D2',
            'created': 'invalid-date'
        }
        response = self.client.post(url, json.dumps(data), content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertIn('Invalid date format', response.json()['error'])


class GenerateExcelTest(TestCase):
    def setUp(self):
        now = timezone.now()

        Robot.objects.create(model='X5', version='LT', created=now - datetime.timedelta(days=2))
        Robot.objects.create(model='X5', version='LT', created=now - datetime.timedelta(days=3))
        Robot.objects.create(model='X7', version='LT', created=now - datetime.timedelta(days=1))
        Robot.objects.create(model='X7', version='Ly', created=now - datetime.timedelta(days=1))
        Robot.objects.create(model='X8', version='gf', created=now - datetime.timedelta(days=4))

    def test_generate_excel(self):
        response = self.client.get(reverse('generate_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        file_stream = response.content
        workbook = load_workbook(filename=io.BytesIO(file_stream))

        expected_sheets = {'Model_X5', 'Model_X7', 'Model_X8'}
        actual_sheets = set(workbook.sheetnames)
        self.assertSetEqual(expected_sheets, actual_sheets)

        sheet_x5 = workbook['Model_X5']
        self.assertEqual(sheet_x5.cell(row=1, column=1).value, 'Model')
        self.assertEqual(sheet_x5.cell(row=2, column=2).value, 'LT')
        self.assertEqual(sheet_x5.cell(row=2, column=3).value, 2)

        sheet_x7 = workbook['Model_X7']
        self.assertEqual(sheet_x7.cell(row=1, column=1).value, 'Model')
        self.assertEqual(sheet_x7.cell(row=2, column=2).value, 'LT')
        self.assertEqual(sheet_x7.cell(row=3, column=2).value, 'Ly')

        sheet_x8 = workbook['Model_X8']
        self.assertEqual(sheet_x8.cell(row=2, column=2).value, 'gf')
        self.assertEqual(sheet_x8.cell(row=2, column=3).value, 1)



class NotifyCustomerOnRobotCreationTest(TestCase):
    def setUp(self):
        self.customer_email = "customer@example.com"
        self.robot_model = "RT"
        self.robot_version = "DT"
        self.robot_serial = f"{self.robot_model}-{self.robot_version}"

        self.customer = Customer.objects.create(email=self.customer_email)

        Order.objects.create(
            customer=self.customer,
            robot_serial=self.robot_serial
        )

    def test_email_sent_when_robot_becomes_available(self):

        Robot.objects.create(
            model=self.robot_model,
            version=self.robot_version,
            created=timezone.now()
        )

        self.assertEqual(len(mail.outbox), 1, "Ожидалось, что отправится одно письмо.")

        email = mail.outbox[0]
        self.assertEqual(email.subject, "Робот теперь в наличии!", "Неверная тема письма.")
        self.assertIn(
            f"модели {self.robot_model}, версии {self.robot_version}",
            email.body,
            "Тело письма не содержит правильную модель и версию робота."
        )
        self.assertEqual(
            email.to,
            [self.customer_email],
            "Письмо отправлено не тому клиенту."
        )
